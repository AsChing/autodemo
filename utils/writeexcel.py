
import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from autodemo.config.public_data import *
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import configparser as cparser

# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read("../config/config.ini", encoding='UTF-8')
name = cf.get("tester", "name")

"""
文件写入数据
"""
class WriteExcel(object):

    def __init__(self, fileName):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝文件至指定目录下
            shutil.copyfile(FILE_PATH, REPORT_FILE_PATH)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self, row_n, result, response):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        """
        try:
            font_GREEN = Font(name='宋体', color="ff10af06", bold=True, size=14)
            font_RED = Font(name='宋体', color="fff10f0f", bold=True, size=14)
            font_Name = Font(name='宋体', color="ff5d0679", bold=True, size=12)
            align = Alignment(horizontal='center', vertical='center')
            # 获数所在行数
            K_n = "K" + str(row_n)
            M_n = "M" + str(row_n)
            # 写入测试结果
            if result == "PASS":
                self.ws.cell(row_n, 11, result)
                self.ws[K_n].font = font_GREEN
            if result == "FAIL":
                self.ws.cell(row_n, 11, result)
                self.ws[K_n].font = font_RED
            # 写入接口返回code
            self.ws.cell(row_n, 12, response)
            # 写入测试人员名称
            self.ws.cell(row_n, 13, name)
            self.ws[K_n].alignment = align
            self.ws[M_n].font = font_Name
            self.ws[M_n].alignment = align
            self.wb.save(self.filename)
            print("----------------写入测试数据成功--------------")
        except Exception as e:
            print("-----------写入失败------------")
            print(e)


if __name__ == '__main__':
    we = WriteExcel(FILE_PATH)
    we.write_data(2, "FAIL", "{'code':400}")
    we.write_data(3, "PASS", "{'code':200}")
